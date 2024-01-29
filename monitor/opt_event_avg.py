import msvcrt
import os
import sys

import pandas as pd
from datetime import datetime

from openpyxl.reader.excel import load_workbook

from util import *
from opt_file_parser import OptFileParser
from MergeData import MergeData


class OptEventAvg(OptFileParser):
    def __init__(self, input_path):
        super().__init__(input_path)
        self.frame = []
        self.stat = {}

    def get_frames(self):
        print(self.cpu_frame_count)
        for frame_index in range(self.cpu_frame_count):
            temp = {
                'frame_index': frame_index,
                'frame_events': []
            }
            for event in self.frame_events[frame_index]:
                event_elapse = get_round_time(event[1] - event[0])
                event_index = event[2]
                temp['frame_events'].append([event_index, event_elapse])
            self.frame.append(temp)

    def merge_elapse(self):
        """合并同一帧中相同事件的耗时"""
        for frame in self.frame:
            main_index_dict = {}
            for event in frame['frame_events']:
                event_index, event_elapse = event[0], event[1]
                if event_index not in main_index_dict.keys():
                    main_index_dict[event_index] = {
                        'elapse': event_elapse
                    }
                else:
                    main_index_dict[event_index]['elapse'] += event_elapse
            # update
            frame['frame_events'] = main_index_dict

    def stat_data(self):
        # frame_elapse = []
        event_elapse = {}
        event_desc_len = len(self.event_descs)
        for i in range(event_desc_len):
            event_elapse[i] = []
        # 平均帧耗时
        avg_frame_elapse = sum(self.cpu_frames_time_ms) / self.cpu_frame_count
        # 提取每帧每事件的耗时
        for frame in self.frame:
            for event_index in frame['frame_events'].keys():  # 事件下标
                event_elapse[event_index].append(frame['frame_events'][event_index]['elapse'])  # 每帧每事件都添加到event_elapse
        # 计算每事件的平均耗时
        for event_index in event_elapse.keys():
            event_name = self.event_descs[event_index].split('(')[0]
            if not event_elapse[event_index]:
                avg_event_elapse = 0
                avg_event_proportion = 0
            else:
                avg_event_elapse = sum(event_elapse[event_index]) / len(event_elapse[event_index])
                avg_event_proportion = (avg_event_elapse / avg_frame_elapse) * 100
            self.stat[event_index] = (event_name, round(avg_event_elapse, 3), round(avg_event_proportion, 3))

    def save(self):
        # output_path = self.input_path[0:-4] + "123" + ".log"
        # recorder = open(output_path, 'w')
        if 'HD' in self.input_path:
            output_file = r'HD_summary.xlsx'
        else:
            output_file = r'BD_summary.xlsx'
        # time_string = datetime.now().strftime("%m-%d")
        data_summary = []
        for key in self.stat.keys():
            if self.stat[key][1] > 0:
                data_summary.append([self.stat[key][0], self.stat[key][1], self.stat[key][2]])
        # 将data_summary中重复的元素，删除
        summary_map = []
        data_summary_tmp = []
        for item in data_summary:
            if item[0] not in summary_map:
                data_summary_tmp.append(item)
                summary_map.append(item[0])
        data_summary = data_summary_tmp
        # 获取文件名
        last_index = self.input_path.rfind('\\')
        second_last_index = self.input_path.rfind('\\', 0, last_index)
        sheet_name = self.input_path[second_last_index + 5:second_last_index + 7] + '-' + self.input_path[second_last_index + 7:second_last_index + 9]
        # 清除已有工作簿
        workbook = load_workbook(output_file)
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        workbook.save(output_file)
        # 保存并写入数据
        df = pd.DataFrame(data_summary, columns=['EventName', 'Avg Event Elapse(ms)', "Avg Event Proportion"])
        if os.path.exists(output_file):
            with pd.ExcelWriter(output_file, mode='a', engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(output_file) as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    def run(self):
        self.read_opt_content()
        self.parse()
        self.get_frames()
        self.merge_elapse()
        self.stat_data()
        self.save()


if __name__ == '__main__':
    is_work = False
    directory = datetime.now().strftime("%Y%m%d")
    directory_path = os.getcwd()
    arguments = sys.argv
    if len(arguments) >= 2:
        directory = arguments[1]
        print(arguments[1])
    # 遍历工作目录下所有文件夹
    for root, dirs, files in os.walk(directory_path):
        for m_dir in dirs:
            # 如果是要搜索的目录
            if directory in m_dir:
                directory = m_dir
                is_work = True
                break
    if is_work:
        file_list = []
        # 获取该目录下文件列表
        for root, dirs, files in os.walk(directory):
            for file in files:
                # 获取文件的完整路径
                file_path = os.path.join(root, file)
                # 将文件路径添加到列表中
                file_list.append(file_path)
        # 处理.opt文件
        for file in file_list:
            print(file)
            if file[-4:] == '.opt':
                xx = OptEventAvg(file)
                xx.run()
        # 处理表格
        xx = MergeData()
        xx.run()
        print('done!')
    else:
        print('Can`t find directory')
